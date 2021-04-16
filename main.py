import json

from openpyxl import load_workbook

workbook = load_workbook(filename="asset_list.xlsx")

sheet = workbook.active
sheet.title = "Monitors"



# create a new entry
def create_log():
    asset_row = len(sheet["A"]) + 1
    print("Please provide details of monitor below")
    model = input("Model: ")
    dimension = int(input("Dimension: "))
    location = input("Location: ")
    condition = input("Condition: ")

    id_cell = "A" + str(asset_row)
    model_cell = "B" + str(asset_row)
    dimension_cell = "C" + str(asset_row)
    location_cell = "D" + str(asset_row)
    condition_cell = "E" + str(asset_row)

    sheet[id_cell] = asset_row
    sheet[model_cell] = model
    sheet[dimension_cell] = dimension
    sheet[location_cell] = location
    sheet[condition_cell] = condition

    workbook.save(filename="asset_list.xlsx")

    return


def read_log(my_row):
    products = {}
    for row in sheet.iter_rows(min_row=my_row,
                               max_row=my_row,
                               values_only=True):
        product_id = row[0]
        product = {
            "model": row[1],
            "dimension": row[2],
            "location": row[3],
            "condition": row[4]
        }
        products[product_id] = product

    print(json.dumps(products))


def get_new():
    pass


def summary():
    pass


def remove():
    pass


def get_number():
    return len(sheet["A"])


def print_rows():
    for row in sheet.iter_rows(values_only=True):
        print(row)


def main():
    read_log()



if __name__=="__main__":
    main()

